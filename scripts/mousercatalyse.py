from openpyxl import Workbook, load_workbook
import datetime, xls2xlsx, argparse, os

ap = argparse.ArgumentParser(description='program for converting mouser shopping carts to spreadsheets importable into catalyse')
ap.add_argument('-i','--input',default='mouser.xls',help='input mouser shopping cart spreadsheet')
ap.add_argument('-o','--output',default='catalyse.xlsx',help='output spreadsheet importable into catalyse')
ap.add_argument('-t','--template',default='template.xlsx',help='empty catalyse template')

args = ap.parse_args()

try:
    pythonpaths = os.environ['PYTHONPATH'].split(';')
except KeyError:
    print('no PYTHONPATH env var configured, only searching for template in cwd.')
    pythonpaths = []
folders = [*pythonpaths,'./']
print('opening catalyse template...')
while True:
    try:
        folder = folders.pop()
        path = os.path.join(folder,args.template)
        print('trying template path:',path)
        wb = load_workbook(path)
        break
    except IndexError:
        print('could not find catalyse template file {}'.format(args.template))
        exit(1)
    except FileNotFoundError:        
        continue
print('template opened.')

ws = wb.active

print('opening mouser shopping cart...')
try:
    wbm = xls2xlsx.XLS2XLSX(args.input).to_xlsx()
except FileNotFoundError:
    print('shopping cart not found')
print('shopping cart opened.')
wsm = wbm.active

# mouser shopping cart format
M_LINO = 'A'
M_MONO = 'B'
M_DESC = 'F'
M_QTY = 'I'
M_PRICE = 'J'

# catalyse spreadsheet format
DESC = 'B'
QTY = 'C'
UNIT = 'D'
DDATE = 'E'
PRICE = 'F'
CURR = 'G'
SUPP = 'H'
SUPP_CODE = 'I'
REF_NUM = 'J'
COMM = 'L'
SUPP_CON = 'M'
TAX = 'O'
BUDG = 'T'

# first row of mouser data
m_line = 10

# first row of catalyse data
c_line = 2
line_delta = 0
while True:
    ml = m_line + line_delta
    cl = c_line + line_delta
    def get_cell(col):
        return wsm['{}{:d}'.format(col,ml)].value
    mlino = get_cell(M_LINO)    
    if not mlino:
        print('end of shopping cart reached.')
        break
    print('adding mouser part # : {}'.format(get_cell(M_MONO)))
    for cc in (
        DESC, QTY, UNIT, DDATE, PRICE, CURR, SUPP, SUPP_CODE, REF_NUM, COMM, SUPP_CON, TAX, BUDG
    ):
        cell = ws['{}{:d}'.format(cc,cl)]        
        if cc == DESC:
            output = get_cell(M_DESC)
        if cc == QTY:
            output = get_cell(M_QTY)
        if cc == UNIT:
            output = 'Piece'
        if cc == DDATE:            
            output = datetime.datetime.today()            
        if cc == PRICE:
            output = float(get_cell(M_PRICE).split()[1])
        if cc == CURR:
            output = get_cell(M_PRICE).split()[0]
        if cc == SUPP:
            output = 'Mouser Electronics'
        if cc == SUPP_CODE:
            output = 'SUP000298'
        if cc == REF_NUM:
            output = get_cell(M_MONO)
        if cc == COMM:
            output = 'Laboratory, measurement, observation and testing equipment'
        if cc == SUPP_CON:
            output = 'Mouser Electronics Info'
        if cc == TAX:
            output = 'TVA CH 7.7'
        if cc == BUDG:
            # "laboratory supplies"
            output = '302000'        
        cell.value = output
    line_delta += 1
print('saving catalyse output spreadsheet.')
wb.save(args.output)
print('spreadsheet saved as {}'.format(args.output))